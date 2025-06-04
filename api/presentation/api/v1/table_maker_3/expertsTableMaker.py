import openpyxl
import openpyxl.styles

class ExpertTableMaker:
    @staticmethod
    def make_excel(metrics, sections, metric_employee_map, emp_name_map, emp_role_map):
        """
        Формирует Excel-файл table_maker_3.xlsx на основании переданных:
          - metrics: список MetricDescription
          - sections: список Section
          - metric_employee_map: {metric_id: employee_id}
          - emp_name_map: {employee_id: "last first surname"}
          - emp_role_map: {employee_id: role_name}
        Итоговый файл сохраняется по пути "api/infrastructure/excel/table_maker_3.xlsx".
        """

        wb = openpyxl.Workbook()
        wb.create_sheet(title='Зоны ответственности', index=0)
        ws = wb['Зоны ответственности']

        bold_italic_font = openpyxl.styles.Font(bold=True, italic=True)
        italic_font = openpyxl.styles.Font(italic=True)

        header_data = ["№", "№", "Показатели", "Имя сотрудника", "Должность"]
        ws.append([
            "Зона ответственности экспертов по вводу сведений в системе оценки деятельности "
            "заведующих кафедрами Московского политехнического университета"
        ])
        title_cell = ws.cell(row=1, column=1)
        title_cell.font = bold_italic_font
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
        ws.append(header_data)

        current_row = 2
        current_category = ""
        current_subname = ""

        for section in sections:
            activity = getattr(section, "activity_name", "")
            subname = getattr(section, "subname", None)

            if activity and current_category != activity:
                current_category = activity
                ws.append([activity])
                cat_cell = ws.cell(row=ws.max_row, column=1)
                cat_cell.font = bold_italic_font
                current_row += 1
                ws.merge_cells(
                    start_row=current_row, start_column=1,
                    end_row=current_row, end_column=5
                )

            if subname:
                if subname != "" and current_subname != subname:
                    current_subname = subname
                    ws.append([subname])
                    sub_cell = ws.cell(row=ws.max_row, column=1)
                    sub_cell.font = italic_font
                    sub_cell.alignment = openpyxl.styles.Alignment(wrap_text=True)
                    current_row += 1
                    ws.merge_cells(
                        start_row=current_row, start_column=1,
                        end_row=current_row, end_column=5
                    )

            for metric in metrics:
                if getattr(metric, "section_id", None) == getattr(section, "id", None):
                    number = getattr(metric, "metric_number", "")
                    subnum = getattr(metric, "metric_subnumber", "") or ""
                    indicator = getattr(metric, "description", "")

                    m_id = getattr(metric, "metric_id", None)
                    employee_id = metric_employee_map.get(m_id)
                    employee_name = emp_name_map.get(employee_id, "")
                    employee_role = emp_role_map.get(employee_id, "")

                    ws.append([number, subnum, indicator, employee_name, employee_role])
                    current_row += 1

        for row_cells in ws.iter_rows(min_row=1, max_row=current_row, min_col=1, max_col=5):
            for cell in row_cells:
                cell.alignment = openpyxl.styles.Alignment(
                    wrap_text=True, horizontal='center', vertical='center'
                )

        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 20

        thin_side = openpyxl.styles.Side(border_style="thin", color="000000")
        for row_cells in ws.iter_rows(min_row=1, max_row=current_row, min_col=1, max_col=5):
            for cell in row_cells:
                cell.border = openpyxl.styles.Border(
                    top=thin_side, left=thin_side,
                    right=thin_side, bottom=thin_side
                )

        # 7. Сохранение файла
        output_path = "api/infrastructure/excel/experts.xlsx"
        wb.save(output_path)
        print(f"Файл успешно сохранен: {output_path}")